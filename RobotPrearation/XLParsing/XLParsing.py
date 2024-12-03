from openpyxl import load_workbook

class   OpenXLReading:

    def readXL(self):
        wb = load_workbook("C:\\Users\\SanjeevaKumarGeejula\\Desktop\\Ramp_GMQA_Resource_Details.xlsx")
        sheet = wb["IPC"]
        print(sheet['A3'].value)
        print(sheet['B5'].value)
        # Iterate through rows and columns
        for rows in sheet.iter_rows(min_row = 1, max_row = 5, min_col = 1, max_col = 3):
            for cell in rows:
                print(cell.value, end=" ")
            print()

xl_reading = OpenXLReading()
xl_reading.readXL()
